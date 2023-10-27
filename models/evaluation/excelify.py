"""
Convert a .json file with the following format for each line:
{"input": "...", "gt": "...", "prediction": "..."}

to an Excel file with the following format for each line:
input | gt | prediction
"""
import json
import os
import openpyxl
from argparse import ArgumentParser
from fuzzywuzzy import fuzz


def main():
    parser = ArgumentParser()
    parser.add_argument('-f', '--file', required=True, nargs="+")
    parser.add_argument('-o', '--output-folder', required=True)
    args = parser.parse_args()

    if not os.path.exists(args.output_folder):
        os.makedirs(args.output_folder)

    for file in args.file:
        if not os.path.exists(file):
            raise ValueError(f"File {file} does not exist.")

    for file in args.file:
        output_file_path = os.path.join(
            args.output_folder, f"{os.path.basename(file).split('.')[0]}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active

        # Get the data from the .json file and convert it into Excel format
        with open(file) as f:
            lines = [json.loads(line) for line in f]

            # Write the header
            ws.append(["Input", "Ground Truth", "Prediction",
                      "EM", "ES"])

            # Make header titles bold and bigger font
            for cell in ws["1:1"]:
                cell.font = openpyxl.styles.Font(bold=True, size=14)
                cell.fill = openpyxl.styles.fills.PatternFill(
                    patternType='solid', fgColor="DDDDDD")

            # Write the data
            for line in lines:

                # Calculate EM and ES
                em = "True" if line["prediction"] == line["gt"] else "False"
                es = fuzz.ratio(line["prediction"], line["gt"]) / 100

                ws.append([
                    readable_json_value(line["input"]),
                    readable_json_value(line["gt"]),
                    readable_json_value(line["prediction"]),
                    em, es])

            # Adjust the column widths and wrap text
            ws.column_dimensions["A"].width = 125
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 50
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

            # Center the text in the cells of columns EM and ES
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal="center")

            # If the value in the fourth column is "True", color the cell green, otherwise red, and allow custom color formatting
            green = openpyxl.styles.colors.Color(rgb='00FF00')
            red = openpyxl.styles.colors.Color(rgb='FF0000')
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    if cell.value == "True":
                        cell.fill = openpyxl.styles.fills.PatternFill(
                            patternType='solid', fgColor=green)
                    else:
                        cell.fill = openpyxl.styles.fills.PatternFill(
                            patternType='solid', fgColor=red)

            # Save the Excel file
            wb.save(output_file_path)


def readable_json_value(input: str) -> str:
    """
    Improve readability of the JSON input by removing EOL tokens and <s> and </s> tokens.
    """
    return input.replace("<EOL>", "\n").replace("<s>", "").replace("</s>", "")


if __name__ == "__main__":
    main()
